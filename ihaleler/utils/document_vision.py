"""
Dosyayı (Excel/PDF/JPG) önce görsele çevirir, sonra Vision LLM ile
teklif cetveli kalemlerini tespit edip yapılandırılmış tablo döndürür.

Kullanım:
    from ihaleler.utils.document_vision import analiz_et_ve_tablo_dondur

    sonuc = analiz_et_ve_tablo_dondur(
        "/path/to/cetvel.pdf",           # veya .xlsx, .jpg
        page_or_sheet_index=0,          # PDF sayfa / Excel sheet
        provider="openai",               # "openai" veya "anthropic"
        ek_talimat="Sadece mal kalemlerini al.",  # isteğe bağlı
    )
    if sonuc["basari"]:
        for satir in sonuc["tablo"]:
            print(satir)  # {"kalem": "...", "birim": "...", "miktar": "...", ...}
    else:
        print(sonuc["hata"])
"""
import base64
import io
import json
import os
import re
from pathlib import Path

# Görsel dönüşüm
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None
try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None
from PIL import Image, ImageDraw, ImageFont
import openpyxl

# Vision API
try:
    from openai import OpenAI
except ImportError:
    OpenAI = None
try:
    from anthropic import Anthropic
except ImportError:
    Anthropic = None


# -----------------------------------------------------------------------------
# 1) DOSYAYI GÖRSELE ÇEVİRME
# -----------------------------------------------------------------------------

def _pdf_to_image_bytes(file_path: str, page_index: int = 0, dpi: int = 150) -> bytes:
    """PDF'in bir sayfasını PNG baytları olarak döndürür. Önce PyMuPDF, yoksa pdf2image."""
    if fitz:
        doc = fitz.open(file_path)
        try:
            page = doc[page_index]
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            return pix.tobytes("png")
        finally:
            doc.close()
    if convert_from_path:
        images = convert_from_path(file_path, first_page=page_index + 1, last_page=page_index + 1, dpi=dpi)
        if images:
            buf = io.BytesIO()
            images[0].save(buf, format="PNG")
            return buf.getvalue()
    raise RuntimeError("PDF'i görsele çevirmek için PyMuPDF veya pdf2image (Poppler) gerekli.")


def _excel_to_image_bytes(file_path: str, sheet_index: int = 0, max_rows: int = 100) -> bytes:
    """Excel sayfasını tablo görünümünde bir PNG olarak çizer (PIL)."""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet_name = wb.sheetnames[sheet_index]
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))[:max_rows]
    finally:
        wb.close()

    if not rows:
        raise ValueError("Excel sayfası boş.")

    # Hücre metinleri; boşları "" yap
    grid = [[str(c) if c is not None else "" for c in row] for row in rows]
    max_cols = max(len(r) for r in grid) if grid else 0
    if max_cols == 0:
        raise ValueError("Excel sayfasında sütun yok.")
    # Sütun genişlikleri (karakter sayısına göre)
    col_widths = []
    for c in range(max_cols):
        lens = [len(str(row[c])) for row in grid if c < len(row)]
        col_widths.append(max(8, min(35, max(lens) if lens else 10)))
    cell_w = 80
    cell_h = 24
    font_size = 11

    width = sum(max(cell_w, min(150, w * 8)) for w in col_widths)
    height = len(grid) * cell_h
    width = min(width, 4000)
    height = min(height, 6000)

    img = Image.new("RGB", (width, height), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", font_size)
    except Exception:
        try:
            font = ImageFont.truetype("C:/Windows/Fonts/arial.ttf", font_size)
        except Exception:
            font = ImageFont.load_default()

    x = 0
    for col_idx, w in enumerate(col_widths):
        cw = max(cell_w, min(150, w * 8))
        for row_idx, row in enumerate(grid):
            y = row_idx * cell_h
            val = (row[col_idx] if col_idx < len(row) else "")[:60]
            draw.rectangle([x, y, x + cw, y + cell_h], outline=(180, 180, 180), fill=(255, 255, 255))
            draw.text((x + 4, y + 2), val, fill=(0, 0, 0), font=font)
        x += cw

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _image_file_to_bytes(file_path: str) -> bytes:
    """JPG/PNG vb. dosyayı olduğu gibi PNG baytı yap (gerekirse dönüştür)."""
    img = Image.open(file_path)
    if img.mode != "RGB":
        img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def file_to_image_bytes(file_path: str, page_or_sheet_index: int = 0) -> bytes:
    """
    Dosyayı (Excel, PDF veya resim) tek bir görselin PNG baytlarına çevirir.

    - PDF: İlk sayfa (veya page_or_sheet_index. sayfa) render edilir.
    - Excel: İlk sayfa (veya page_or_sheet_index. sheet) tablo olarak çizilir.
    - .jpg, .jpeg, .png: Dosya doğrudan PNG baytına dönüştürülür.

    Returns:
        PNG formatında görsel baytları.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Dosya bulunamadı: {file_path}")
    ext = path.suffix.lower()
    if ext == ".pdf":
        return _pdf_to_image_bytes(file_path, page_index=page_or_sheet_index)
    if ext in (".xlsx", ".xls"):
        return _excel_to_image_bytes(file_path, sheet_index=page_or_sheet_index)
    if ext in (".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"):
        return _image_file_to_bytes(file_path)
    raise ValueError(f"Desteklenmeyen format: {ext}. Kullanılabilir: .pdf, .xlsx, .xls, .jpg, .jpeg, .png")


# -----------------------------------------------------------------------------
# 2) VISION LLM İLE ANALİZ VE YAPILANDIRILMIŞ TABLO
# -----------------------------------------------------------------------------

VISION_SYSTEM_PROMPT = """Sen bir teklif cetveli / birim fiyat listesi analiz uzmanısın.
Sana bir sayfa görseli verilecek. Bu sayfaya bir insan gibi bak; teklif cetvelindeki kalemleri (ürün/hizmet adları), birimleri ve miktarları tespit et.
Sayfa yapısı ne olursa olsun (tablo, serbest metin, tarama, Excel çıktısı vb.) verileri yapılandırılmış bir tablo olarak döndür.
Sütun isimlerine bağımlı kalma; görsel yerleşimden (layout) anlam çıkar. Örneğin: kalem açıklaması, birim (adet, m², kg vb.), miktar, birim fiyat, toplam vb. alanları yerleşime göre eşleştir.

Yanıtını SADECE aşağıdaki JSON formatında ver. Başka açıklama ekleme.
Format:
{"tablo": [{"kalem": "...", "birim": "...", "miktar": "...", "birim_fiyat": "...", "toplam": "...", "aciklama": "..."}, ...]}

Eksik sütunlar olabilir; gördüğün sütunları key olarak kullan (örn. kalem, birim, miktar, birim_fiyat, toplam, aciklama, sira_no).
Her satır bir kalem olsun. Başlık satırlarını tabloya ekleme."""


def _get_api_keys():
    try:
        from django.conf import settings
        return {
            "openai": getattr(settings, "OPENAI_API_KEY", None) or os.environ.get("OPENAI_API_KEY"),
            "anthropic": getattr(settings, "ANTHROPIC_API_KEY", None) or os.environ.get("ANTHROPIC_API_KEY"),
        }
    except Exception:
        return {
            "openai": os.environ.get("OPENAI_API_KEY"),
            "anthropic": os.environ.get("ANTHROPIC_API_KEY"),
        }


def _call_openai_vision(image_base64: str, user_prompt: str, api_key: str) -> str:
    """OpenAI GPT-4o ile görsel analiz."""
    if not OpenAI:
        raise RuntimeError("openai paketi yüklü değil: pip install openai")
    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": VISION_SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_prompt},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_base64}"}},
                ],
            },
        ],
        max_tokens=4096,
    )
    return (response.choices[0].message.content or "").strip()


def _call_anthropic_vision(image_base64: str, user_prompt: str, api_key: str) -> str:
    """Anthropic Claude ile görsel analiz."""
    if not Anthropic:
        raise RuntimeError("anthropic paketi yüklü değil: pip install anthropic")
    client = Anthropic(api_key=api_key)
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        system=VISION_SYSTEM_PROMPT,
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_prompt},
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": "image/png",
                            "data": image_base64,
                        },
                    },
                ],
            },
        ],
    )
    return (response.content[0].text if response.content else "").strip()


def _parse_table_from_response(text: str) -> list:
    """Yanıt metninden JSON tablo çıkarır; bulamazsa boş liste."""
    text = text.strip()
    # JSON bloğu ara (```json ... ``` veya doğrudan { ... })
    json_match = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text)
    if json_match:
        text = json_match.group(1)
    else:
        obj_match = re.search(r"\{[\s\S]*\}", text)
        if obj_match:
            text = obj_match.group(0)
    try:
        data = json.loads(text)
        if isinstance(data, dict) and "tablo" in data:
            return data["tablo"] if isinstance(data["tablo"], list) else []
        if isinstance(data, list):
            return data
        return []
    except json.JSONDecodeError:
        return []


def analiz_et_ve_tablo_dondur(
    file_path: str,
    page_or_sheet_index: int = 0,
    provider: str = "openai",
    ek_talimat: str = "",
) -> dict:
    """
    Dosyayı (Excel/PDF/JPG) önce görsele çevirir, Vision LLM'e gönderir;
    teklif cetveli kalemlerini tespit edip yapılandırılmış tablo döndürür.

    Args:
        file_path: Dosya yolu (.pdf, .xlsx, .xls, .jpg, .jpeg, .png).
        page_or_sheet_index: PDF için sayfa indeksi (0=ilk), Excel için sheet indeksi.
        provider: "openai" veya "anthropic". Hangi Vision API kullanılacak.
        ek_talimat: LLM'e ek metin talimatı (isteğe bağlı).

    Returns:
        {
            "basari": True/False,
            "tablo": [ {"kalem": "...", "birim": "...", "miktar": "...", ...}, ... ],
            "ham_yanit": "modelin döndürdüğü ham metin",
            "hata": "varsa hata mesajı"
        }
    """
    result = {"basari": False, "tablo": [], "ham_yanit": "", "hata": None}
    user_prompt = "Bu sayfadaki teklif cetveli / birim fiyat kalemlerini tespit et ve yapılandırılmış tablo olarak döndür."
    if ek_talimat:
        user_prompt += "\n\nEk talimat: " + ek_talimat

    try:
        image_bytes = file_to_image_bytes(file_path, page_or_sheet_index=page_or_sheet_index)
    except Exception as e:
        result["hata"] = f"Dosya görsele çevrilemedi: {e}"
        return result

    b64 = base64.b64encode(image_bytes).decode("utf-8")
    keys = _get_api_keys()

    if provider == "openai":
        api_key = keys["openai"]
        if not api_key:
            result["hata"] = "OPENAI_API_KEY bulunamadı (.env veya settings)"
            return result
        try:
            result["ham_yanit"] = _call_openai_vision(b64, user_prompt, api_key)
        except Exception as e:
            result["hata"] = f"OpenAI Vision hatası: {e}"
            return result
    elif provider == "anthropic":
        api_key = keys["anthropic"]
        if not api_key:
            result["hata"] = "ANTHROPIC_API_KEY bulunamadı (.env veya settings)"
            return result
        try:
            result["ham_yanit"] = _call_anthropic_vision(b64, user_prompt, api_key)
        except Exception as e:
            result["hata"] = f"Anthropic Vision hatası: {e}"
            return result
    else:
        result["hata"] = f"Bilinmeyen provider: {provider}. 'openai' veya 'anthropic' kullanın."
        return result

    result["tablo"] = _parse_table_from_response(result["ham_yanit"])
    result["basari"] = True
    return result
