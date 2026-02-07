from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Q
from django.urls import reverse
from django.http import HttpResponse
from decimal import Decimal

from django.contrib.auth.models import User
from .models import UrunKutuphanesi, Kalem, Ihale, Hastane, Arac, AracKullanimKaydi


def _ihale_queryset(request, is_dogrudan_temin):
    """İhale listesini filtre ve sıralamaya göre döndürür."""
    qs = Ihale.objects.filter(is_dogrudan_temin=is_dogrudan_temin).select_related("hastane").prefetch_related("kalemler").order_by("-tarih")
    q = request.GET.get("q", "").strip()
    if q:
        qs = qs.filter(
            Q(ihale_adi__icontains=q) | Q(hastane__ad__icontains=q) | Q(kalemler__urun_adi__icontains=q)
        ).distinct()
    il = request.GET.get("il", "").strip()
    if il:
        qs = qs.filter(il__iexact=il)
    tur = request.GET.get("tur", "").strip()
    if tur:
        qs = qs.filter(tur=tur)
    durum = request.GET.get("durum", "").strip()
    if durum:
        qs = qs.filter(durum=durum)
    sirala = request.GET.get("sirala", "sisteme_yeni")
    if sirala == "sisteme_eski":
        qs = qs.order_by("tarih")
    elif sirala == "ihale_yeniden_eski":
        qs = qs.order_by("-tarih")
    elif sirala == "ihale_eskiden_yeni":
        qs = qs.order_by("tarih")
    else:
        qs = qs.order_by("-olusturulma_tarihi")
    return qs


# =========================
# ANA SAYFA
# =========================
def anasayfa(request):
    # Açık ihaleler (is_dogrudan_temin=False, durum=Acik)
    toplam_ihale = Ihale.objects.filter(is_dogrudan_temin=False, durum='Acik').count()
    # Açık doğrudan teminler
    toplam_dogrudan = Ihale.objects.filter(is_dogrudan_temin=True, durum='Acik').count()
    # Aktif kurum: hem açık ihale hem açık doğrudan temini olan hastaneler (tekil)
    ihale_hastane_ids = set(
        Ihale.objects.filter(is_dogrudan_temin=False, durum='Acik').values_list('hastane_id', flat=True).distinct()
    )
    dt_hastane_ids = set(
        Ihale.objects.filter(is_dogrudan_temin=True, durum='Acik').values_list('hastane_id', flat=True).distinct()
    )
    aktif_kurum_sayisi = len(ihale_hastane_ids & dt_hastane_ids)
    return render(request, "ihaleler/index.html", {
        "toplam_ihale": toplam_ihale,
        "toplam_dogrudan": toplam_dogrudan,
        "aktif_kurum_sayisi": aktif_kurum_sayisi,
    })


# =========================
# İHALE & DOĞRUDAN TEMİN
# =========================
def ihale_listesi(request):
    ihaleler = _ihale_queryset(request, is_dogrudan_temin=False)
    return render(request, "ihaleler/ihaleler.html", {"ihaleler": ihaleler, "blur_content": not request.user.is_authenticated})


def dogrudan_temin_listesi(request):
    ihaleler = _ihale_queryset(request, is_dogrudan_temin=True)
    return render(request, "ihaleler/dogrudan_temin.html", {"ihaleler": ihaleler, "blur_content": not request.user.is_authenticated})


def liste_filtre(request):
    """Ana sayfadaki Mal/Yapım/Hizmet/Açık tıklanınca: tek sayfada hem ihale hem doğrudan temin listesi (filtreli)."""
    tur = request.GET.get("tur", "").strip()
    durum = request.GET.get("durum", "").strip()
    qs_ihale = Ihale.objects.filter(is_dogrudan_temin=False).select_related("hastane").prefetch_related("kalemler").order_by("-tarih")
    qs_dt = Ihale.objects.filter(is_dogrudan_temin=True).select_related("hastane").prefetch_related("kalemler").order_by("-tarih")
    if tur:
        qs_ihale = qs_ihale.filter(tur=tur)
        qs_dt = qs_dt.filter(tur=tur)
    if durum:
        qs_ihale = qs_ihale.filter(durum=durum)
        qs_dt = qs_dt.filter(durum=durum)
    return render(request, "ihaleler/liste_filtre.html", {
        "ihaleler": qs_ihale,
        "dogrudan_teminler": qs_dt,
        "filtre_tur": tur,
        "filtre_durum": durum,
        "blur_content": not request.user.is_authenticated,
    })


@login_required
def ihale_detay(request, pk):
    """İhale / doğrudan temin dosyasının tam sayfa detayı; kalemler ve geçmiş linkleri."""
    ihale = get_object_or_404(Ihale.objects.prefetch_related("kalemler"), pk=pk)
    return render(request, "ihaleler/ihale_detay.html", {"ihale": ihale})


@login_required
def ihale_detay_excel_indir(request, pk):
    """İhale detay sayfasındaki tüm verileri Excel olarak indirir."""
    import openpyxl
    from openpyxl.styles import Font

    ihale = get_object_or_404(Ihale.objects.prefetch_related("kalemler"), pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detay"

    # Dosya bilgileri
    ws.append(["İhale / İKN No", ihale.ihale_no or ""])
    ws.append(["İhale Adı", ihale.ihale_adi or ""])
    ws.append(["Tarih", ihale.tarih.strftime("%d.%m.%Y") if ihale.tarih else ""])
    ws.append(["Hastane / Kurum", ihale.hastane.ad if ihale.hastane else ""])
    ws.append(["İl / İlçe", "{} {}".format(ihale.il or "", ihale.ilce or "").strip()])
    ws.append(["Tür", ihale.tur or ""])
    ws.append(["Durum", ihale.get_durum_display() if ihale.durum else ""])
    ws.append(["Doğrudan Temin", "Evet" if ihale.is_dogrudan_temin else "Hayır"])
    ws.append(["Toplam Teklif Bedeli (₺)", str(ihale.toplam_teklif_bedeli or 0)])
    ws.append(["Bizim Teklif (₺)", str(ihale.bizim_teklif or "")])
    ws.append(["Kazanan Firma", ihale.kazanan_firma or ""])
    ws.append(["Kazanan Fiyat (₺)", str(ihale.kazanan_fiyat or "")])
    ws.append(["Yükleyen", ihale.olusturan_kullanici.username if ihale.olusturan_kullanici else ""])
    ws.append([])

    # Kalemler tablosu
    ws.append(["Ürün Adı", "Miktar", "Birim", "Birim Fiyat (₺)", "Toplam (₺)"])
    for kalem in ihale.kalemler.all():
        ws.append([
            kalem.urun_adi or "",
            float(kalem.adet) if kalem.adet is not None else 0,
            kalem.birim or "",
            float(kalem.birim_fiyat) if kalem.birim_fiyat is not None else 0,
            float(kalem.toplam_fiyat) if kalem.toplam_fiyat is not None else 0,
        ])

    # Başlık satırı kalın
    for row in range(1, 16):
        ws.cell(row=row, column=1).font = Font(bold=True)
    for c in range(1, 6):
        ws.cell(row=16, column=c).font = Font(bold=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    dosya_adi = "ihale_{}_detay.xlsx".format((ihale.ihale_no or str(pk)).replace("/", "-").replace(" ", "_"))
    response["Content-Disposition"] = 'attachment; filename="{}"'.format(dosya_adi)
    wb.save(response)
    return response


@login_required
def kalem_gecmis(request, pk):
    """Kalemin ürün geçmişi: aynı ürünün geçtiği ihaleler, görseller, faturalar, fiyatlar."""
    kalem = get_object_or_404(Kalem, pk=pk)
    # Aynı ürün adına sahip tüm kalemler (bu ihale + diğer ihaleler)
    gecmis = Kalem.objects.filter(urun_adi__iexact=kalem.urun_adi).select_related("ihale").order_by("-ihale__tarih")
    return render(request, "ihaleler/kalem_gecmis.html", {"kalem": kalem, "gecmis": gecmis})


# =========================
# DOSYA & İŞLEMLER
# =========================
def dosya_yukleme(request):
    from .services import dosya_yukleme as dosya_yukleme_handler
    return dosya_yukleme_handler(request)


@login_required
def ihale_sil(request, pk):
    """İhaleyi veya doğrudan temin kaydını siler, listeye yönlendirir."""
    ihale = get_object_or_404(Ihale, pk=pk)
    is_dogrudan = ihale.is_dogrudan_temin
    ihale.delete()
    from django.contrib import messages
    messages.success(request, "Kayıt silindi.")
    if is_dogrudan:
        return redirect(reverse("dogrudan_temin_listesi"))
    return redirect(reverse("ihale_listesi"))


def ihale_excel_indir(request):
    return render(request, "ihaleler/index.html")


# =========================
# TOPLU İŞLEMLER
# =========================
@login_required
def toplu_fiyat_guncelle(request):
    from django.contrib import messages
    if request.method != "POST":
        return redirect(request.META.get("HTTP_REFERER", reverse("ihale_listesi")))
    ids = request.POST.getlist("kalem_id[]") or request.POST.getlist("kalem_id")
    fiyatlar = request.POST.getlist("fiyat[]") or request.POST.getlist("fiyat")
    if not ids:
        messages.info(request, "Kaydedilecek kalem bulunamadı.")
        return redirect(request.META.get("HTTP_REFERER", reverse("ihale_listesi")))
    for i, kid in enumerate(ids):
        if not kid:
            continue
        try:
            kalem = Kalem.objects.get(pk=int(kid))
        except (Kalem.DoesNotExist, ValueError):
            continue
        try:
            bf = Decimal(str(fiyatlar[i] if i < len(fiyatlar) else "0").replace(",", "."))
        except Exception:
            bf = Decimal("0")
        kalem.birim_fiyat = bf
        kalem.toplam_fiyat = kalem.adet * bf
        kalem.save()
    messages.success(request, "Fiyatlar güncellendi.")
    return redirect(request.META.get("HTTP_REFERER", reverse("ihale_listesi")))


@login_required
def kalem_ekle(request, pk):
    """İhaleye manuel kalem ekler. POST: urun_adi, adet, birim, birim_fiyat, toplam_fiyat."""
    ihale = get_object_or_404(Ihale, pk=pk)
    if request.method != "POST":
        return redirect("ihale_listesi")
    urun_adi = (request.POST.get("urun_adi") or "").strip()[:255]
    if not urun_adi:
        from django.contrib import messages
        messages.warning(request, "Ürün adı zorunludur.")
        return redirect(request.META.get("HTTP_REFERER", reverse("ihale_listesi")))
    try:
        adet = Decimal(str(request.POST.get("adet", "1")).replace(",", "."))
    except Exception:
        adet = Decimal("1")
    birim = (request.POST.get("birim") or "Adet")[:50]
    try:
        birim_fiyat = Decimal(str(request.POST.get("birim_fiyat", "0")).replace(",", "."))
    except Exception:
        birim_fiyat = Decimal("0")
    try:
        toplam_fiyat = Decimal(str(request.POST.get("toplam_fiyat", "0")).replace(",", "."))
    except Exception:
        toplam_fiyat = adet * birim_fiyat if birim_fiyat else Decimal("0")
    Kalem.objects.create(
        ihale=ihale,
        urun_adi=urun_adi,
        adet=adet,
        birim=birim,
        birim_fiyat=birim_fiyat,
        toplam_fiyat=toplam_fiyat,
    )
    from django.contrib import messages
    messages.success(request, "Kalem eklendi.")
    detay_url = reverse("ihale_detay", args=[pk]) + "#kalem-ekle"
    return redirect(detay_url)


@login_required
def kalem_gorsel_yukle(request, pk):
    """Kalem için ürün görseli yükler."""
    kalem = get_object_or_404(Kalem, pk=pk)
    if request.method == "POST" and request.FILES.get("kalem_gorsel"):
        kalem.kalem_gorsel = request.FILES["kalem_gorsel"]
        kalem.save()
        from django.contrib import messages
        messages.success(request, "Kalem görseli yüklendi.")
    return redirect(request.META.get("HTTP_REFERER", reverse("ihale_listesi")))


@login_required
def ihale_verilen_teklif_yukle(request, pk):
    """İhaleye 'Verilen Teklif' (ıslak imzalı çıktı) dosyası yükler."""
    ihale = get_object_or_404(Ihale, pk=pk)
    if request.method == "POST" and request.FILES.get("verilen_teklif_dosya"):
        ihale.verilen_teklif_dosya = request.FILES["verilen_teklif_dosya"]
        ihale.save()
        from django.contrib import messages
        messages.success(request, "Verilen teklif dosyası yüklendi.")
    return redirect(request.META.get("HTTP_REFERER", reverse("ihale_listesi")))


# =========================
# PROFİL & ANALİZ (giriş zorunlu)
# =========================
@login_required
def profilim(request):
    return render(request, "ihaleler/profilim.html")


@login_required
def analiz_sayfasi(request):
    if request.method == "POST" and request.POST.get("sonuc_guncelle"):
        ihale_id = request.POST.get("ihale_id")
        if ihale_id:
            ihale = get_object_or_404(Ihale, pk=ihale_id)
            ihale.kazanan_firma = (request.POST.get("kazanan_firma") or "").strip() or None
            try:
                v = request.POST.get("kazanan_fiyat")
                ihale.kazanan_fiyat = Decimal(v.replace(",", ".")) if v else None
            except Exception:
                ihale.kazanan_fiyat = None
            try:
                v = request.POST.get("bizim_teklif")
                ihale.bizim_teklif = Decimal(v.replace(",", ".")) if v else None
            except Exception:
                ihale.bizim_teklif = None
            if not ihale.bizim_teklif and ihale.toplam_teklif_bedeli:
                ihale.bizim_teklif = ihale.toplam_teklif_bedeli
            durum = request.POST.get("durum")
            if durum in dict(Ihale.DURUM_CHOICES):
                ihale.durum = durum
            ihale.save()
        return redirect("analiz")

    hastane_verileri = Hastane.objects.annotate(
        toplam=Count("ihale_set"),
    ).order_by("-toplam")

    en_cok_alinanlar = (
        Kalem.objects.values("urun_adi")
        .annotate(adet_sum=Sum("adet"))
        .order_by("-adet_sum")[:10]
    )

    ihaleler = Ihale.objects.all().order_by("-tarih")[:50]
    sonuc_ihaleler = []
    for i in ihaleler:
        bizim = i.bizim_teklif or i.toplam_teklif_bedeli or Decimal("0")
        kazanan = i.kazanan_fiyat or Decimal("0")
        kazandik = i.durum == "Kazandik"
        net_kar = None
        fark_yuzde = None
        if kazandik and i.toplam_maliyet is not None:
            satis = i.satis_fiyati or i.toplam_teklif_bedeli or Decimal("0")
            net_kar = (satis or Decimal("0")) - (i.toplam_maliyet or Decimal("0"))
        elif i.durum == "Kaybettik" and kazanan and bizim:
            fark_yuzde = ((bizim - kazanan) / bizim * 100) if bizim else None
        sonuc_ihaleler.append({
            "ihale": i,
            "bizim_teklif": bizim,
            "kazanan_fiyat": kazanan,
            "kazandik": kazandik,
            "net_kar": net_kar,
            "fark_yuzde": fark_yuzde,
        })

    kaybedilen = [
        s for s in sonuc_ihaleler
        if s["ihale"].durum == "Kaybettik" and s["kazanan_fiyat"] and s["bizim_teklif"]
    ]
    ortalama_fark_yuzde = None
    if kaybedilen:
        toplam = sum((s["fark_yuzde"] or 0) for s in kaybedilen)
        ortalama_fark_yuzde = toplam / len(kaybedilen)

    return render(request, "ihaleler/analiz.html", {
        "hastane_verileri": hastane_verileri,
        "en_cok_alinanlar": en_cok_alinanlar,
        "sonuc_ihaleler": sonuc_ihaleler,
        "kaybedilen_ihaleler": kaybedilen,
        "ortalama_fark_yuzde": ortalama_fark_yuzde,
        "durum_choices": Ihale.DURUM_CHOICES,
    })


# =========================
# MESAİ (giriş zorunlu)
# =========================
@login_required
def mesailerim_view(request):
    return render(request, "ihaleler/mesailerim.html")


# =========================
# ARAÇLAR (giriş zorunlu)
# =========================
@login_required
def arabalar_view(request):
    from django.utils import timezone
    from django.contrib import messages

    if request.method == "POST":
        # Yeni araç tanımla
        if request.POST.get("arac_ekle"):
            plaka = (request.POST.get("plaka") or "").strip().upper()
            marka = (request.POST.get("marka") or "").strip()
            if not plaka or not marka:
                messages.warning(request, "Plaka ve Marka/Model zorunludur.")
            else:
                try:
                    mevcut_km = int(request.POST.get("mevcut_km") or 0)
                    if mevcut_km < 0:
                        mevcut_km = 0
                except (ValueError, TypeError):
                    mevcut_km = 0
                personel_id = request.POST.get("personel")
                zimmet = User.objects.get(pk=int(personel_id)) if personel_id else None
                arac = Arac.objects.create(
                    plaka=plaka,
                    marka_model=marka,
                    mevcut_km=mevcut_km,
                    zimmetli_personel=zimmet,
                )
                if request.FILES.get("arac_foto"):
                    arac.arac_foto = request.FILES["arac_foto"]
                    arac.save()
                messages.success(request, f"Araç {plaka} kaydedildi.")
            return redirect(reverse("arabalar"))

        # Araç al (kullanıma al)
        if request.POST.get("arac_al"):
            arac_id = request.POST.get("arac_id")
            personel_id = request.POST.get("personel_id")
            try:
                arac = Arac.objects.get(pk=int(arac_id))
                personel = User.objects.get(pk=int(personel_id))
            except (Arac.DoesNotExist, User.DoesNotExist, ValueError, TypeError):
                messages.warning(request, "Geçersiz araç veya personel.")
                return redirect(reverse("arabalar"))
            if arac.su_an_kullanimda:
                messages.warning(request, "Bu araç zaten kullanımda.")
                return redirect(reverse("arabalar"))
            try:
                baslangic_km = int(request.POST.get("baslangic_km") or 0)
                if baslangic_km < 0:
                    baslangic_km = 0
            except (ValueError, TypeError):
                baslangic_km = arac.mevcut_km
            ihale_id = request.POST.get("ihale_id")
            ihale = Ihale.objects.get(pk=int(ihale_id)) if ihale_id else None
            alis_str = request.POST.get("alis_tarihi")
            alis_tarihi = None
            if alis_str:
                try:
                    from datetime import datetime
                    alis_tarihi = timezone.make_aware(datetime.strptime(alis_str.replace("T", " ")[:19], "%Y-%m-%d %H:%M"))
                except Exception:
                    pass
            AracKullanimKaydi.objects.create(
                arac=arac,
                personel=personel,
                ihale=ihale,
                baslangic_km=baslangic_km,
                alis_tarihi=alis_tarihi or timezone.now(),
            )
            messages.success(request, f"Araç {arac.plaka} üzerinize alındı.")
            return redirect(reverse("arabalar"))

        # Araç teslim et
        if request.POST.get("arac_teslim"):
            arac_id = request.POST.get("arac_id")
            try:
                arac = Arac.objects.get(pk=int(arac_id))
            except (Arac.DoesNotExist, ValueError, TypeError):
                messages.warning(request, "Araç bulunamadı.")
                return redirect(reverse("arabalar"))
            kayit = arac.su_an_kullanimda
            if not kayit:
                messages.warning(request, "Bu araç zaten teslim edilmiş.")
                return redirect(reverse("arabalar"))
            try:
                bitis_km = int(request.POST.get("bitis_km") or 0)
                if bitis_km < 0:
                    bitis_km = 0
            except (ValueError, TypeError):
                bitis_km = kayit.baslangic_km
            teslim_str = request.POST.get("teslim_tarihi")
            teslim_tarihi = None
            if teslim_str:
                try:
                    from datetime import datetime
                    teslim_tarihi = timezone.make_aware(datetime.strptime(teslim_str.replace("T", " ")[:19], "%Y-%m-%d %H:%M"))
                except Exception:
                    pass
            kayit.bitis_km = bitis_km
            kayit.teslim_tarihi = teslim_tarihi or timezone.now()
            kayit.teslim_notu = (request.POST.get("teslim_notu") or "").strip() or None
            if request.FILES.get("teslim_gorseli"):
                kayit.teslim_gorseli = request.FILES["teslim_gorseli"]
            kayit.save()
            arac.mevcut_km = bitis_km
            arac.save(update_fields=["mevcut_km"])
            messages.success(request, f"Araç {arac.plaka} teslim kaydı tamamlandı.")
            return redirect(reverse("arabalar"))

    # GET: liste + filtre
    araclar = Arac.objects.all().prefetch_related("kullanim_gecmisi").order_by("plaka")
    q = request.GET.get("q", "").strip()
    if q:
        araclar = araclar.filter(
            Q(plaka__icontains=q) | Q(marka_model__icontains=q)
        )
    durum = request.GET.get("durum", "").strip()
    if durum == "musait":
        araclar = [a for a in araclar if not a.su_an_kullanimda]
    elif durum == "kullanimda":
        araclar = [a for a in araclar if a.su_an_kullanimda]
    else:
        araclar = list(araclar)
    personeller = User.objects.filter(is_active=True).order_by("username")
    ihaleler = Ihale.objects.all().order_by("-tarih")[:200]
    return render(request, "ihaleler/arabalar.html", {
        "araclar": araclar,
        "personeller": personeller,
        "ihaleler": ihaleler,
    })


# =========================
# ÜRÜN KATALOG & GEÇMİŞ ANALİZ
# =========================
@login_required
def urun_katalog_listesi(request):
    """Ürün Katalog listesi: Katalogdaki tüm ürünler, kaç ihalede kullanıldığı."""
    urunler = UrunKutuphanesi.objects.annotate(
        kalem_sayisi=Count("kalem_set"),
        ihale_sayisi=Count("kalem_set__ihale", distinct=True),
    ).order_by("-kalem_sayisi")
    return render(request, "ihaleler/urun_katalog.html", {"urunler": urunler})


@login_required
def urun_gecmis_analizi(request, pk):
    """Ürün Geçmiş Analizi: Alış/satış, hangi ihalelerde kullanıldığı, toplam kar."""
    urun = get_object_or_404(UrunKutuphanesi, pk=pk)
    kalemler = Kalem.objects.filter(kutuphane_urunu=urun).select_related("ihale").order_by("-ihale__tarih")

    toplam_alis = Decimal("0")
    toplam_satis = Decimal("0")
    satirlar = []
    for k in kalemler:
        alis = (k.alinan_fiyat or k.maliyet_birim_fiyat * k.adet if k.maliyet_birim_fiyat else Decimal("0"))
        satis = (k.satis_fiyati or k.toplam_fiyat or Decimal("0"))
        kar = satis - alis
        toplam_alis += alis
        toplam_satis += satis
        satirlar.append({
            "kalem": k,
            "ihale": k.ihale,
            "alis": alis,
            "satis": satis,
            "kar": kar,
        })
    toplam_kar = toplam_satis - toplam_alis

    return render(request, "ihaleler/urun_gecmis_analiz.html", {
        "urun": urun,
        "satirlar": satirlar,
        "toplam_alis": toplam_alis,
        "toplam_satis": toplam_satis,
        "toplam_kar": toplam_kar,
        "kalem_sayisi": len(satirlar),
    })


# =========================
# İLETİŞİM
# =========================
# Büyükşehirler başta: İstanbul, Ankara, İzmir önce; sonra diğer iller
IL_SECIM = [
    "İstanbul", "Ankara", "İzmir", "Bursa", "Antalya", "Adana", "Kocaeli", "Kayseri",
    "Gaziantep", "Mersin", "Konya", "Diyarbakır", "Şanlıurfa", "Balıkesir", "Samsun",
    "Trabzon", "Manisa", "Hatay", "Aydın", "Denizli", "Muğla", "Malatya", "Mardin", "Tekirdağ",
]
GOREV_SECIM = [
    "Satın Alma", "Teknik Şartname", "İdari İşler", "Mali İşler", "Başhekim",
    "İdari İşler Müdürü", "Teknik Hizmetler", "Bilgi İşlem", "İnsan Kaynakları", "Diğer",
]


def iletisim_view(request):
    iller = list(IL_SECIM)  # Sıra büyükşehirler başta
    ilce_list = []
    hastaneler = []
    if request.GET.get("il"):
        il = request.GET.get("il")
        ilce_list = list(
            Ihale.objects.filter(il=il).values_list("ilce", flat=True).distinct()
        )
        ilce_list = [x for x in ilce_list if x]
        ilce_list.sort()
    if request.GET.get("ilce"):
        ilce = request.GET.get("ilce")
        il = request.GET.get("il", "")
        hastaneler = list(
            Hastane.objects.filter(
                ihale__il=il, ihale__ilce=ilce
            ).distinct().values_list("id", "ad")
        )
    elif request.GET.get("il"):
        il = request.GET.get("il")
        hastaneler = list(
            Hastane.objects.filter(ihale__il=il).distinct().values_list("id", "ad")
        )
    return render(request, "ihaleler/iletisim.html", {
        "iller": iller,
        "ilce_list": ilce_list,
        "hastaneler": hastaneler,
        "gorevler": GOREV_SECIM,
    })
